from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from functools import lru_cache
from pathlib import Path

APP_DIR = Path(__file__).resolve().parent
WORKSPACE_DIR = APP_DIR
PROJECT_DATA_DIR = APP_DIR / "project_data"
BPS_RESOURCES_DIR = APP_DIR / "resources" / "bps"
BPS_HS_FILE = BPS_RESOURCES_DIR / "hscode_clean_BPS.xlsx"
LEGACY_BPS_HS_FILE = APP_DIR.parent / "hscode_clean_BPS.xlsx"
ML_RESOURCES_DIR = APP_DIR / "resources" / "ml"
ML_MODEL_PATH = ML_RESOURCES_DIR / "model_rf_gravity.pkl"
ML_ENCODER_PATH = ML_RESOURCES_DIR / "encoder_rf_gravity.pkl"

BPS_YEAR_RANGE = tuple(range(date.today().year - 10, date.today().year))
EXPECTED_BPS_YEARS = BPS_YEAR_RANGE
EXPECTED_BPS_YEAR_LABEL = f"{EXPECTED_BPS_YEARS[0]}-{EXPECTED_BPS_YEARS[-1]}"
DEFAULT_BPS_YEAR = str(EXPECTED_BPS_YEARS[-1])


def _resolve_bps_hs_file():
    if BPS_HS_FILE.exists():
        return BPS_HS_FILE
    return LEGACY_BPS_HS_FILE


@lru_cache(maxsize=1)
def _get_default_bps_start_hs():
    hs_path = _resolve_bps_hs_file()
    if not hs_path.exists():
        return ""

    try:
        from openpyxl import load_workbook

        workbook = load_workbook(hs_path, read_only=True, data_only=True)
        worksheet = workbook.active
        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        header_cells = [str(value).strip().lower() if value is not None else "" for value in header_row]
        if "hs_code" not in header_cells:
            workbook.close()
            return ""

        hs_index = header_cells.index("hs_code")
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if hs_index >= len(row):
                continue
            value = row[hs_index]
            if value is None:
                continue
            text = str(value).strip()
            if text:
                workbook.close()
                return text

        workbook.close()
    except Exception:
        return ""

    return ""

WORKSPACE_ORDER = ("ssinas", "bps", "trademap", "ml")
WORKSPACE_META = {
    "ssinas": {
        "page_title": "siinas.kemenperin.go.id Workspace",
        "nav_title": "siinas.kemenperin.go.id",
        "nav_subtitle": "Portal Internal",
        "short": "SI",
        "sidebar_copy": (
            "Masuk ke workspace untuk scraper yang butuh session aktif, token manual, "
            "dan kontrol halaman yang lebih terarah."
        ),
        "hero_title": "Workspace Scraping siinas.kemenperin.go.id",
        "hero_copy": (
            "Area khusus untuk scraping portal siinas.kemenperin.go.id. Fokusnya ke flow token manual, "
            "range halaman, dan pemantauan hasil insert ke database."
        ),
        "hero_badge": "Portal workflow",
        "hero_note": (
            "Tampilan dipisah supaya tool portal internal tidak bercampur dengan runner BPS, "
            "jadi lebih enak dipantau saat kerja harian."
        ),
        "section_heading": "Tool siinas.kemenperin.go.id",
        "section_caption": (
            "Gunakan halaman ini kalau kamu mau menjalankan scraper yang butuh login ke "
            "portal siinas.kemenperin.go.id dan interaksi browser."
        ),
        "jobs_heading": "Run siinas.kemenperin.go.id Terakhir",
        "jobs_caption": (
            "Daftar ini hanya menampilkan job dari workspace siinas.kemenperin.go.id supaya log dan statusnya "
            "tidak bercampur dengan job BPS."
        ),
        "architecture_notes": (
            {
                "title": "Login Manual",
                "text": (
                    "Untuk VPS, user cukup paste PHPSESSID atau full cookie header dari browser pribadi. "
                    "Server tidak perlu membuka Chrome GUI."
                ),
            },
            {
                "title": "Range Halaman",
                "text": (
                    "Input start page dan end page membantu kamu membatasi scraping ke rentang "
                    "yang memang sedang dibutuhkan."
                ),
            },
            {
                "title": "Stop Aman",
                "text": (
                    "Job siinas.kemenperin.go.id bisa dihentikan dari dashboard tanpa perlu mematikan "
                    "server Flask, baik memakai token manual maupun mode browser lokal."
                ),
            },
        ),
    },
    "bps": {
        "page_title": "BPS Workspace",
        "nav_title": "BPS",
        "nav_subtitle": "Trade API",
        "short": "BP",
        "sidebar_copy": (
            "Masuk ke workspace BPS untuk runner trade gabungan dengan pilihan tahun dan "
            "jenis data export atau import."
        ),
        "hero_title": "Workspace Scraping BPS",
        "hero_copy": (
            "Runner BPS digabung ke satu tempat. Kamu tinggal pilih tahun dan flow data, "
            "lalu jalankan dari dashboard atau terminal terpisah."
        ),
        "hero_badge": "Unified runner",
        "hero_note": (
            "Semua variasi BPS sekarang tidak menumpuk di dashboard utama. Satu workspace ini "
            f"cukup untuk semua kebutuhan trade API {EXPECTED_BPS_YEAR_LABEL}."
        ),
        "section_heading": "Tool BPS",
        "section_caption": (
            "Halaman ini dikhususkan untuk runner BPS gabungan, jadi pilihan tahun dan export "
            "atau import bisa dikelola dari satu form saja."
        ),
        "jobs_heading": "Run BPS Terakhir",
        "jobs_caption": (
            "Daftar ini hanya menampilkan job dari workspace BPS agar mudah memantau proses "
            "background dan terminal terpisah."
        ),
        "architecture_notes": (
            {
                "title": "Satu Runner BPS",
                "text": (
                    "Semua variasi BPS disatukan ke dalam `project_data/scrapapibps.py`, "
                    "jadi tahun dan jenis data dipilih dari form."
                ),
            },
            {
                "title": "Mode Run Ganda",
                "text": (
                    "Runner bisa dijalankan sebagai background job dari dashboard atau dibuka "
                    "ke PowerShell terpisah saat kamu butuh paralel."
                ),
            },
            {
                "title": "Lebih Mudah Dirawat",
                "text": (
                    "Kalau ada perubahan logika API atau database, cukup ubah satu script dan "
                    "semua variasi tahun langsung ikut."
                ),
            },
        ),
    },
    "trademap": {
        "page_title": "TradeMap Workspace",
        "nav_title": "TradeMap",
        "nav_subtitle": "Folder Parser",
        "short": "TM",
        "sidebar_copy": (
            "Workspace ini disiapkan untuk import folder Excel TradeMap dari direktori lokal, "
            "tanpa scraping browser. Fokusnya parsing file dan insert ke database."
        ),
        "hero_title": "Workspace Parsing TradeMap",
        "hero_copy": (
            "Halaman ini dipakai untuk mengolah folder berisi file Excel atau HTML-XLS hasil "
            "export TradeMap, lalu memetakan isinya ke Model 1 atau Model 2 sesuai kebutuhan database."
        ),
        "hero_badge": "Folder-based parser",
        "hero_note": (
            "Scraping browser TradeMap memang belum kita pindahkan ke dashboard ini, tapi parser "
            "foldernya sudah disiapkan supaya file hasil export bisa langsung diproses ke database."
        ),
        "section_heading": "Tool TradeMap",
        "section_caption": (
            "Halaman ini dipakai untuk upload folder, deteksi format file TradeMap, parsing, "
            "dan insert ke tabel target yang sesuai dengan mode parser."
        ),
        "jobs_heading": "Run TradeMap Terakhir",
        "jobs_caption": (
            "Semua aktivitas import folder TradeMap akan muncul di bagian ini agar terpisah "
            "dari SSINAS dan BPS."
        ),
        "architecture_notes": (
            {
                "title": "Dua Format Input",
                "text": (
                    "`scrapindov3.py` membaca file HTML yang disimpan sebagai `.xls`, sedangkan "
                    "`dataolah.py` dan `datanew.py` membaca file Excel template biasa `.xls/.xlsx`."
                ),
            },
            {
                "title": "Dua Model Output",
                "text": (
                    "Mode Negara-Mitra masuk ke `data_perdagangan_full_v3`, sedangkan Mode Negara-ALL "
                    "masuk ke `data_perdagangan`."
                ),
            },
            {
                "title": "Auto Create Tabel Tambahan",
                "text": (
                    "Tabel `data_perdagangan` ikut dibuat saat konfigurasi database, dan parser tetap "
                    "bisa memastikannya tersedia dari file `data_perdagangan (2).sql`."
                ),
            },
        ),
    },
    "ml": {
        "page_title": "ML Predictor Workspace",
        "nav_title": "ML Predictor",
        "nav_subtitle": "Trade Forecast",
        "short": "ML",
        "sidebar_copy": (
            "Workspace ini dipakai untuk prediksi machine learning berbasis model gravity trade. "
            "Fokus utamanya sekarang adalah upload dataset manual dengan format yang sesuai model."
        ),
        "hero_title": "Workspace Prediksi Machine Learning",
        "hero_copy": (
            "Halaman ini menyatukan model gravity trade ke dalam dashboard utama. "
            "Prediksi difokuskan ke upload dataset manual agar user tidak bergantung pada tabel referensi GDP dan Growth di database."
        ),
        "hero_badge": "Model Random Forest",
        "hero_note": (
            "Model ini diambil dari implementasi Flask lama lalu dipindahkan ke project utama, "
            "supaya workflow scraping dan prediksi ada di satu dashboard dengan format input yang lebih terarah."
        ),
        "section_heading": "Tool Prediksi",
        "section_caption": (
            "Workspace ini fokus ke prediksi ekspor berbasis model machine learning, dengan input manual yang mengikuti format dataset model."
        ),
        "jobs_heading": "Riwayat Prediksi",
        "jobs_caption": (
            "Halaman ini tidak memakai job background seperti scraper. Hasil prediksi akan ditampilkan langsung di workspace ini."
        ),
        "architecture_notes": (
            {
                "title": "Manual Upload Only",
                "text": (
                    "Mode prediksi sekarang difokuskan ke upload dataset manual penuh, supaya user tidak perlu menunggu tabel referensi GDP dan Growth tersedia di database."
                ),
            },
            {
                "title": "Model Gravity",
                "text": (
                    "Prediksi memakai fitur Reporter, Partner, HS4, Year, GDP, dan GDP Growth yang sama seperti aplikasi ML yang sudah jadi sebelumnya."
                ),
            },
            {
                "title": "Validasi Format",
                "text": (
                    "Workspace ini memberi tahu format dataset yang didukung, menyiapkan template contoh, dan menampilkan preview hasil agar user mudah mengikuti polanya."
                ),
            },
        ),
    },
}


@dataclass(frozen=True)
class FormField:
    name: str
    label: str
    field_type: str = "text"
    default: str = ""
    min_value: str | None = None
    placeholder: str = ""
    help_text: str = ""
    options: tuple[tuple[str, str], ...] = ()
    accept: str = ""
    allow_multiple: bool = False
    directory: bool = False


@dataclass(frozen=True)
class ToolDefinition:
    key: str
    title: str
    subtitle: str
    category: str
    workspace: str
    description: str
    run_kind: str
    command: str
    status_label: str
    status_tone: str
    note: str
    ready: bool = True
    form_fields: tuple[FormField, ...] = ()
    script_path: Path | None = None


def normalize_workspace(workspace: str | None):
    if workspace in WORKSPACE_META:
        return workspace
    return WORKSPACE_ORDER[0]


def _ssinas_tool():
    return ToolDefinition(
        key="ssinas",
        title="siinas.kemenperin.go.id Session Token",
        subtitle="Manual token + Flask",
        category="Portal Internal",
        workspace="ssinas",
        description=(
            "Scraper siinas.kemenperin.go.id memakai session aktif dari browser user, "
            "sehingga server VPS tidak perlu membuka browser GUI."
        ),
        run_kind="internal",
        command="Paste PHPSESSID/cookie aktif, lalu jalankan scraping dari dashboard.",
        status_label="Ready",
        status_tone="success",
        note="Cocok untuk VPS karena tidak wajib membuka Chrome di server.",
        ready=True,
        form_fields=(
            FormField(
                name="auth_mode",
                label="Auth Mode",
                field_type="select",
                default="manual_token",
                help_text="Gunakan token manual untuk VPS. Browser manual hanya untuk lokal dengan GUI.",
                options=(
                    ("manual_token", "Manual PHPSESSID / Cookie"),
                    ("browser_manual", "Browser Manual Lokal"),
                ),
            ),
            FormField(
                name="phpsessid",
                label="PHPSESSID",
                field_type="password",
                placeholder="Paste value PHPSESSID dari browser",
                help_text="Token tidak disimpan di payload job; hanya dipakai saat job berjalan.",
            ),
            FormField(
                name="cookie_header",
                label="Full Cookie Header",
                field_type="text",
                placeholder="Opsional: PHPSESSID=...; cookie_lain=...",
                help_text="Gunakan jika portal membutuhkan cookie tambahan selain PHPSESSID.",
            ),
            FormField(
                name="start_page",
                label="Start Page",
                field_type="number",
                default="1",
                min_value="1",
                help_text="Halaman awal yang ingin diproses.",
            ),
            FormField(
                name="end_page",
                label="End Page",
                field_type="number",
                default="5",
                min_value="1",
                help_text="Halaman terakhir yang akan diproses.",
            ),
        ),
    )


def _bps_tool():
    script_path = PROJECT_DATA_DIR / "scrapapibps.py"
    hs_path = _resolve_bps_hs_file()
    default_start_hs = _get_default_bps_start_hs()

    missing_parts = []
    if not script_path.exists():
        missing_parts.append("script scrapapibps.py")
    if not hs_path.exists():
        missing_parts.append("file resources/bps/hscode_clean_BPS.xlsx")

    ready = not missing_parts
    status_label = "Ready" if ready else "Perlu setup"
    status_tone = "success" if ready else "warning"
    note = (
        f"Satu runner BPS untuk 10 tahun terakhir ({EXPECTED_BPS_YEAR_LABEL}) dan jenis export/import."
        if ready
        else "Konfigurasi belum lengkap: " + ", ".join(missing_parts) + "."
    )

    return ToolDefinition(
        key="bps-trade",
        title="BPS Trade Unified",
        subtitle=f"API Perdagangan BPS {EXPECTED_BPS_YEAR_LABEL}",
        category="BPS Trade API",
        workspace="bps",
        description=(
            "Satu script BPS gabungan. Pilih tahun dan jenis data langsung dari dashboard, "
            "tanpa perlu memelihara banyak file scraper yang logiknya sama."
        ),
        run_kind="script",
        command=f'python "{script_path}" --year <tahun> --flow <export|import>',
        status_label=status_label,
        status_tone=status_tone,
        note=note,
        ready=ready,
        script_path=script_path,
        form_fields=(
            FormField(
                name="year",
                label="Tahun",
                field_type="select",
                default=DEFAULT_BPS_YEAR,
                help_text=f"Pilih salah satu dari 10 tahun terakhir: {EXPECTED_BPS_YEAR_LABEL}.",
                options=tuple((str(year), str(year)) for year in EXPECTED_BPS_YEARS),
            ),
            FormField(
                name="flow",
                label="Jenis Data",
                field_type="select",
                default="export",
                help_text="Pilih jenis perdagangan yang akan dijalankan.",
                options=(
                    ("export", "Export"),
                    ("import", "Import"),
                ),
            ),
            FormField(
                name="start_from_hs",
                label="Resume Dari HS",
                default=default_start_hs,
                placeholder="Kosongkan untuk mulai dari awal",
                help_text=(
                    "Default mengikuti HS code pertama di file sumber. "
                    "Ganti nilainya kalau ingin melanjutkan dari titik tertentu."
                ),
            ),
        ),
    )


def _trademap_tool():
    script_path = PROJECT_DATA_DIR / "trademap_folder_parser.py"
    ready = script_path.exists()
    status_label = "Ready" if ready else "Perlu setup"
    note = (
        "Mode Negara-Mitra masuk ke `data_perdagangan_full_v3`, sedangkan Mode Negara-ALL masuk ke "
        "`data_perdagangan` yang sekarang ikut dibuat saat setup database."
        if ready
        else "Script parser TradeMap belum ditemukan di project_data/trademap_folder_parser.py."
    )

    return ToolDefinition(
        key="trademap-folder-parser",
        title="TradeMap Folder Import",
        subtitle="Negara-Mitra + Negara-ALL",
        category="TradeMap Parser",
        workspace="trademap",
        description=(
            "Import satu folder TradeMap, pilih mode Negara-Mitra atau Negara-ALL, lalu kirim hasil "
            "parsingnya ke tabel tujuan yang sesuai tanpa perlu scraping browser."
        ),
        run_kind="script",
        command="Upload folder TradeMap lalu jalankan parser ke database dari dashboard.",
        status_label=status_label,
        status_tone="success" if ready else "warning",
        note=note,
        ready=ready,
        script_path=script_path,
        form_fields=(
            FormField(
                name="parser_mode",
                label="Mode Parser",
                field_type="select",
                default="model_1_default",
                help_text="Negara-Mitra mengarah ke `data_perdagangan_full_v3`, sedangkan Negara-ALL mengarah ke `data_perdagangan`.",
                options=(
                    ("model_1_default", "Negara-Mitra"),
                    ("model_2_negara_all", "Negara-ALL"),
                ),
            ),
            FormField(
                name="status",
                label="Status Data",
                field_type="select",
                default="export",
                help_text="Pilih apakah folder berisi data export atau import.",
                options=(
                    ("export", "Export"),
                    ("import", "Import"),
                ),
            ),
            FormField(
                name="parser_profile",
                label="Profil Parser",
                field_type="select",
                default="auto",
                help_text="Pakai auto agar parser memilih format file yang cocok untuk mode yang sedang dipakai.",
                options=(
                    ("auto", "Auto Detect"),
                    ("html", "HTML TradeMap"),
                    ("excel", "Excel Template"),
                ),
            ),
            FormField(
                name="folder_upload",
                label="Upload Folder TradeMap",
                field_type="file",
                default="",
                accept=".xls,.xlsx",
                allow_multiple=True,
                directory=True,
                help_text="Pilih folder yang berisi file TradeMap. Browser akan mengirim semua file di dalam folder itu.",
            ),
            FormField(
                name="folder_source",
                label="Folder Source",
                field_type="hidden",
                default="",
            ),
        ),
    )


def _ml_tool():
    ready = ML_MODEL_PATH.exists() and ML_ENCODER_PATH.exists()
    status_label = "Ready" if ready else "Perlu setup"
    status_tone = "success" if ready else "warning"
    note = (
        "Model Random Forest gravity trade aktif. Gunakan halaman workspace untuk upload dataset manual yang sesuai format model."
        if ready
        else "File model atau encoder ML belum ditemukan di resources/ml."
    )

    return ToolDefinition(
        key="ml-gravity-predictor",
        title="Trade Gravity Predictor",
        subtitle="Random Forest + Encoder",
        category="Machine Learning",
        workspace="ml",
        description=(
            "Prediksi potensi nilai ekspor menggunakan model machine learning yang sudah diintegrasikan ke dashboard utama."
        ),
        run_kind="internal",
        command="Gunakan halaman workspace ML Predictor untuk memilih mode tbtrade atau upload dataset manual.",
        status_label=status_label,
        status_tone=status_tone,
        note=note,
        ready=ready,
    )


def discover_tools():
    return [_ssinas_tool(), _bps_tool(), _trademap_tool(), _ml_tool()]


def get_tool(tool_key: str):
    for tool in discover_tools():
        if tool.key == tool_key:
            return tool
    return None


def _build_tool_groups(tools):
    groups = []
    for category in sorted({tool.category for tool in tools}):
        groups.append(
            {
                "name": category,
                "tools": [tool for tool in tools if tool.category == category],
            }
        )
    return groups


def _build_nav_items(tools, current_workspace):
    items = []
    for workspace in WORKSPACE_ORDER:
        meta = WORKSPACE_META[workspace]
        workspace_tools = [tool for tool in tools if tool.workspace == workspace]
        items.append(
            {
                "key": workspace,
                "title": meta["nav_title"],
                "subtitle": meta["nav_subtitle"],
                "short": meta["short"],
                "count": len(workspace_tools),
                "ready": len([tool for tool in workspace_tools if tool.ready]),
                "is_active": workspace == current_workspace,
            }
        )
    return items


def _build_stats(workspace, tools):
    ready_count = len([tool for tool in tools if tool.ready])

    if workspace == "bps":
        return (
            {
                "label": "Runner",
                "value": str(len(tools)),
                "caption": "Runner gabungan BPS yang aktif di dashboard.",
            },
            {
                "label": "Tahun",
                "value": str(len(tuple(EXPECTED_BPS_YEARS))),
                "caption": f"Pilihan tahun otomatis mengikuti 10 tahun terakhir: {EXPECTED_BPS_YEAR_LABEL}.",
            },
            {
                "label": "Jenis Data",
                "value": "2",
                "caption": "Export dan import ditangani oleh runner yang sama.",
            },
            {
                "label": "Siap Jalan",
                "value": str(ready_count),
                "caption": "Runner yang konfigurasi dasarnya sudah lengkap.",
            },
        )

    if workspace == "trademap":
        return (
            {
                "label": "Tool Draft",
                "value": str(len(tools)),
                "caption": "Parser TradeMap yang sedang disiapkan untuk aplikasi.",
            },
            {
                "label": "Model Parser",
                "value": "2",
                "caption": "Mode Negara-Mitra dan Negara-ALL tersedia di workspace ini.",
            },
            {
                "label": "Target Tabel",
                "value": "2",
                "caption": "`data_perdagangan_full_v3` dan `data_perdagangan` dipakai sesuai mode parser.",
            },
            {
                "label": "Siap Jalan",
                "value": str(ready_count),
                "caption": "Parser yang sudah bisa dijalankan dari dashboard.",
            },
        )

    if workspace == "ml":
        return (
            {
                "label": "Model",
                "value": "1",
                "caption": "Satu model gravity Random Forest aktif di workspace ini.",
            },
            {
                "label": "Mode Input",
                "value": "1",
                "caption": "Prediksi difokuskan ke upload dataset manual penuh yang mengikuti format model.",
            },
            {
                "label": "Fitur Model",
                "value": "8",
                "caption": "Reporter, Partner, HS4, Year, GDP, dan Growth dipakai saat inferensi.",
            },
            {
                "label": "Siap Jalan",
                "value": str(ready_count),
                "caption": "Tool prediksi yang aset modelnya sudah lengkap di project utama.",
            },
        )

    form_count = len(tools[0].form_fields) if tools else 0
    return (
        {
            "label": "Tool Aktif",
            "value": str(len(tools)),
            "caption": "Tool siinas.kemenperin.go.id yang tersedia di workspace ini.",
        },
        {
            "label": "Siap Jalan",
            "value": str(ready_count),
            "caption": "Tool dengan konfigurasi dasar yang sudah siap dipakai.",
        },
        {
            "label": "Kontrol Form",
            "value": str(form_count),
            "caption": "Input utama untuk mengatur rentang scraping halaman.",
        },
        {
            "label": "Mode Login",
            "value": "Manual",
            "caption": "Browser dibuka untuk login dan CAPTCHA saat proses dimulai.",
        },
    )


def build_dashboard_context(current_workspace=None):
    tools = discover_tools()
    current_workspace = normalize_workspace(current_workspace)
    workspace_meta = WORKSPACE_META[current_workspace]
    workspace_tools = [tool for tool in tools if tool.workspace == current_workspace]

    return {
        "page_title": workspace_meta["page_title"],
        "current_section": current_workspace,
        "current_workspace": workspace_meta,
        "nav_items": _build_nav_items(tools, current_workspace),
        "hero_title": workspace_meta["hero_title"],
        "hero_copy": workspace_meta["hero_copy"],
        "hero_badge": workspace_meta["hero_badge"],
        "hero_note": workspace_meta["hero_note"],
        "tool_groups": _build_tool_groups(workspace_tools),
        "stats": _build_stats(current_workspace, workspace_tools),
        "missing_slots": (),
        "architecture_notes": workspace_meta["architecture_notes"],
        "section_heading": workspace_meta["section_heading"],
        "section_caption": workspace_meta["section_caption"],
        "jobs_heading": workspace_meta["jobs_heading"],
        "jobs_caption": workspace_meta["jobs_caption"],
    }
